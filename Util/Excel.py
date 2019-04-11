import xlrd
from openpyxl import load_workbook

class OpenrationExcel:
    def __init__(self,file_name=None,sheet_id=None):
        if file_name:
            self.file_name = file_name
            self.sheet_id = sheet_id
        else:
            self.file_name =r'D:\pycharm\Interface_new_2019-3-7_mooc\data1\mooc_data.xlsx'
            self.sheet_id = 0
        self.data = self.get_data()


    """
    获取sheets的内容
    """
    def get_data(self):
        try:
            data = xlrd.open_workbook(self.file_name)
            tables = data.sheets()[self.sheet_id]
        except:
            print("exce文件正在被打开")
        return tables

    """
    获取单元格的行数
    """
    def get_lines(self):
        tables = self.data
        return tables.nrows

    """
    获取某一个单元格内容
    """
    def get_cell_value(self,row,col):
        return self.data.cell_value(row,col)

    """
    写入数据
    """
    def write_value(self,row,col,value):
        '''
        :param row: 行
        :param col: 列
        :param value:写入的内容
        :return: None
        '''
        try:
            read_data = load_workbook(self.file_name)
            read_data1 = read_data.active
            read_data1.cell(row,col,value)
            read_data.save(self.file_name)
        except:
            print("excel正在被打开")

    """根据对应的case_id找到对应行的内容"""
    def get_rows_data(self,case_id):
        row_num = self.get_row_num(case_id)
        rows_data = self.get_row_values(row_num)
        return rows_data

    """根据对应的case_id找到对应的行号"""
    def get_row_num(self,case_id):
        num = 0
        cols_data = self.get_cols_data()
        for col_data in cols_data:
            if case_id in col_data:
                return num
            num = num +1

    """根据行号找到该行的内容"""
    def get_row_values(self,row):
        tables = self.data
        row_data = tables.row_values(row)
        return row_data

    """获取某一列的内容"""
    def get_cols_data(self,col_id=None):
        if col_id !=None:
            cols = self.data.col_values(col_id)
        else:
            cols = self.data.col_values(0)
        return cols




if __name__ == '__main__':
    oe = OpenrationExcel()
    # print(oe.get_lines())
    # print(oe.get_cell_value(1,0))
    #oe.write_value(2,11,"pass")
    #print(oe.get_cols_data(1))
    #print(oe.get_rows_data("web_002"))
    value = '"code":10001004,"msg":"\u9a8c\u8bc1\u7801\u65e0\u6548"'
    oe.write_value(1,12,value)

